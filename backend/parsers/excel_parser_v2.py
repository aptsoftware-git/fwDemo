"""
Improved Excel parser for FRS Data Management System.
Handles multi-row headers and merged cells correctly.
"""
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pandas as pd
import numpy as np
from difflib import SequenceMatcher
import jellyfish
import openpyxl

# Import config only when not in test mode
try:
    from config import SHEET_TYPES, SHEET_TYPE_ORDER
except ImportError:
    # Fallback for testing
    SHEET_TYPES = {
        'APPX_A_BVEH': 'Appx A (B veh)',
        'APPX_A_CVEH': 'Appx A (C veh)',
        'ARMT': 'armt',
        'SA': 'SA',
        'INST': 'INST',
        'CBRN': 'CBRN'
    }
    SHEET_TYPE_ORDER = ['APPX_A_BVEH', 'APPX_A_CVEH', 'ARMT', 'SA', 'INST', 'CBRN']


def calculate_similarity(str1: str, str2: str) -> float:
    """Calculate similarity between two strings using multiple methods."""
    # Normalize strings
    s1 = ''.join(c.lower() for c in str1 if c.isalnum() or c.isspace()).strip()
    s2 = ''.join(c.lower() for c in str2 if c.isalnum() or c.isspace()).strip()
    
    # Sequence matcher
    seq_similarity = SequenceMatcher(None, s1, s2).ratio()
    
    # Soundex
    try:
        soundex_match = 1.0 if jellyfish.soundex(s1) == jellyfish.soundex(s2) else 0.0
    except:
        soundex_match = 0.0
    
    return (seq_similarity * 0.7) + (soundex_match * 0.3)


def match_sheet_name(actual_name: str, expected_names: Dict[str, str], threshold: float = 0.6) -> Optional[str]:
    """Match an actual sheet name to expected sheet types using fuzzy matching."""
    best_match = None
    best_score = threshold
    
    for sheet_type, expected_name in expected_names.items():
        similarity = calculate_similarity(actual_name, expected_name)
        if similarity > best_score:
            best_score = similarity
            best_match = sheet_type
    
    return best_match


def detect_header_rows(ws: openpyxl.worksheet.worksheet.Worksheet, max_scan_rows: int = 10) -> Tuple[int, List[str]]:
    """
    Detect where headers end and data begins, and extract column names.
    
    Args:
        ws: openpyxl worksheet
        max_scan_rows: Maximum rows to scan for headers
    
    Returns:
        Tuple of (data_start_row, column_names)
    """
    # Scan first rows to find header structure
    rows_data = []
    for row_idx in range(1, max_scan_rows + 1):
        row_values = []
        for col_idx in range(1, 25):  # Scan first 25 columns
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value if cell.value is not None else ""
            row_values.append(str(value).strip())
        rows_data.append(row_values)
    
    # Strategy: Find header rows containing header keywords
    header_keywords = ['auth', 'held', 'fmc', 'category', 'make', 'type', 'ser', 'remarks', 
                       'dependency', 'mua', 'total', 'nomenclature', 'eqpt', 'nmc']
    
    # Identify header rows (before markers like (i), (ii))
    header_row_indices = []
    marker_row_idx = None
    
    for idx, row_values in enumerate(rows_data):
        row_text = ' '.join(row_values).lower()
        keyword_count = sum(1 for kw in header_keywords if kw in row_text)
        
        # Check if this is a marker row (contains (i), (ii), etc.)
        has_markers = any(v.startswith('(') and v.endswith(')') and len(v) <= 5 for v in row_values if v)
        
        if keyword_count >= 2 and not has_markers:
            header_row_indices.append(idx)
        elif has_markers and marker_row_idx is None:
            marker_row_idx = idx
            break  # Stop scanning after marker row
    
    if not header_row_indices:
        # Fallback: assume rows 1-2 are headers
        header_row_indices = [1, 2]
    
    # Extract column names by checking each column across header rows
    column_names = []
    header_rows = [rows_data[i] for i in header_row_indices if i < len(rows_data)]
    
    if not header_rows:
        return 5, []  # Default fallback
    
    # Find max columns
    max_cols = max(len(row) for row in header_rows)
    
    # For each column, collect non-empty header values
    for col_idx in range(max_cols):
        col_parts = []
        for header_row in header_rows:
            if col_idx < len(header_row):
                value = header_row[col_idx].strip()
                # Skip empty, markers, and generic words
                if value and not (value.startswith('(') and value.endswith(')')):
                    # Skip if it's a parent header that will be repeated (like "Dependency")
                    if value.lower() not in ['dependency', 'nmc due to', 'under']:
                        col_parts.append(value)
        
        # Pick the best column name
        if col_parts:
            # Use the most specific (usually last) non-empty value
            column_names.append(col_parts[-1])
        else:
            # No name found, generate generic
            column_names.append(f"Column_{col_idx + 1}")
    
    # Data starts after marker row or after last header
    if marker_row_idx is not None:
        data_start_row = marker_row_idx + 2  # After markers, skip one more row
    else:
        data_start_row = max(header_row_indices) + 2
    
    # Skip category rows (all caps, first column only)
    while data_start_row <= max_scan_rows:
        if data_start_row - 1 < len(rows_data):
            row_values = rows_data[data_start_row - 1]
            first_val = row_values[0] if row_values else ""
            # Check if it's a category header row
            if first_val and first_val.isupper() and not first_val.isdigit():
                # Check if rest of row is mostly empty
                non_empty_count = sum(1 for v in row_values[1:5] if v)
                if non_empty_count < 2:
                    data_start_row += 1
                    continue
        break
    
    return data_start_row, column_names


def parse_sheet_with_detection(file_path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Parse a single sheet with automatic header detection.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to parse
    
    Returns:
        pandas DataFrame with properly named columns
    """
    # Load workbook to analyze structure
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
    
    ws = wb[sheet_name]
    
    # Detect headers and data start
    data_start_row, column_names = detect_header_rows(ws)
    
    wb.close()
    
    # Read data using pandas, skipping header rows
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, skiprows=data_start_row - 1)
    
    # Set column names
    if column_names and len(column_names) <= len(df.columns):
        df.columns = column_names + [f"Extra_{i}" for i in range(len(column_names), len(df.columns))]
    else:
        # Fallback: use detected names as much as possible
        df.columns = (column_names + [f"Column_{i}" for i in range(len(column_names), len(df.columns))])[:len(df.columns)]
    
    # Remove completely empty rows
    df = df.dropna(how='all')
    
    # Remove rows that are category separators or marker rows
    def is_skip_row(row):
        first_val = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
        
        # Skip marker rows like (i), (ii), (iii)
        if first_val.startswith('(') and first_val.endswith(')') and len(first_val) <= 5:
            return True
        
        # Skip category header rows (all caps, mostly empty)
        if first_val and first_val.isupper() and not first_val.isdigit():
            non_empty = row.iloc[1:5].notna().sum()
            if non_empty < 2:
                return True
        
        # Skip rows where first column contains header-like text
        if any(keyword in first_val.lower() for keyword in ['category', 'make', 'eqpt', 'ser no']):
            return True
            
        return False
    
    # Filter out skip rows
    mask = ~df.apply(is_skip_row, axis=1)
    df = df[mask]
    
    # Reset index
    df = df.reset_index(drop=True)
    
    return df


def parse_excel_file(file_path: str, dictionary_mapping: Dict[str, List[str]] = None) -> Dict[str, pd.DataFrame]:
    """
    Parse an Excel file and extract data from all sheets with improved header detection.
    
    Args:
        file_path: Path to Excel file (.xls or .xlsx)
        dictionary_mapping: Column mapping from dictionary parser (optional, currently not used)
    
    Returns:
        Dictionary mapping sheet_type to pandas DataFrame
    
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file format is unsupported
        Exception: If parsing fails
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    file_ext = file_path.suffix.lower()
    
    if file_ext not in ['.xls', '.xlsx']:
        raise ValueError(f"Unsupported file format: {file_ext}. Expected .xls or .xlsx")
    
    try:
        # Get sheet names
        if file_ext == '.xlsx':
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
        else:
            excel_file = pd.ExcelFile(file_path, engine='xlrd')
        
        sheet_names = excel_file.sheet_names
        excel_file.close()
        
        # Dictionary to store parsed sheets
        parsed_sheets = {}
        matched_types = set()
        
        # Try fuzzy matching for each sheet
        for sheet_name in sheet_names:
            matched_type = match_sheet_name(sheet_name, SHEET_TYPES)
            
            if matched_type and matched_type not in matched_types:
                # Parse the sheet with automatic header detection
                df = parse_sheet_with_detection(file_path, sheet_name)
                
                parsed_sheets[matched_type] = df
                matched_types.add(matched_type)
        
        # Fallback: Use sequence-based matching for unmatched types
        if len(parsed_sheets) < 6:
            unmatched_types = [st for st in SHEET_TYPE_ORDER if st not in matched_types]
            unmatched_sheets = [sn for sn in sheet_names if not any(
                match_sheet_name(sn, SHEET_TYPES) == mt for mt in matched_types
            )]
            
            for sheet_type, sheet_name in zip(unmatched_types, unmatched_sheets):
                if sheet_type and sheet_name:
                    df = parse_sheet_with_detection(file_path, sheet_name)
                    parsed_sheets[sheet_type] = df
        
        return parsed_sheets
    
    except Exception as e:
        raise Exception(f"Failed to parse Excel file {file_path}: {str(e)}")


def get_sheet_summary(df: pd.DataFrame) -> Dict:
    """Get summary statistics for a DataFrame."""
    return {
        'row_count': len(df),
        'column_count': len(df.columns),
        'columns': list(df.columns),
        'has_null': df.isnull().any().any(),
        'null_counts': df.isnull().sum().to_dict()
    }


def validate_excel_structure(file_path: str) -> Tuple[bool, List[str]]:
    """Validate that an Excel file has the expected structure."""
    issues = []
    
    try:
        file_path = Path(file_path)
        
        if not file_path.exists():
            return False, ["File does not exist"]
        
        if file_path.suffix.lower() not in ['.xls', '.xlsx']:
            return False, ["Invalid file format (expected .xls or .xlsx)"]
        
        # Try to parse the file
        parsed_sheets = parse_excel_file(str(file_path))
        
        # Check if we have enough sheets
        if len(parsed_sheets) == 0:
            issues.append("No sheets could be parsed")
        elif len(parsed_sheets) < 4:  # At minimum, expect 4 sheets
            issues.append(f"Only {len(parsed_sheets)} sheets parsed, expected at least 4")
        
        # Check each sheet for basic validity
        for sheet_type, df in parsed_sheets.items():
            if df.empty:
                issues.append(f"Sheet {sheet_type} is empty")
            elif len(df.columns) < 5:
                issues.append(f"Sheet {sheet_type} has too few columns ({len(df.columns)})")
        
        is_valid = len(issues) == 0
        
        return is_valid, issues
    
    except Exception as e:
        return False, [f"Validation error: {str(e)}"]


# For testing
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        test_file = sys.argv[1]
    else:
        test_file = "../data/FRS_Sample/Nov/Fmn A Nov/Nov 2025.xlsx"
    
    try:
        print(f"Parsing: {test_file}")
        parsed = parse_excel_file(test_file)
        
        print(f"\nSuccessfully parsed {len(parsed)} sheets:")
        for sheet_type, df in parsed.items():
            print(f"\n{'='*80}")
            print(f"{sheet_type}:")
            print(f"{'='*80}")
            summary = get_sheet_summary(df)
            print(f"  Rows: {summary['row_count']}, Columns: {summary['column_count']}")
            print(f"  Column names: {summary['columns']}")
            print(f"\n  First 3 rows:")
            print(df.head(3))
    
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
